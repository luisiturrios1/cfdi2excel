import os
from threading import Thread
from tkinter import Button, Frame, Label, Tk
from tkinter.filedialog import askdirectory
from tkinter.messagebox import askokcancel, showinfo
from tkinter.ttk import Progressbar

from lxml import etree
from openpyxl import Workbook

NSMAP = {
    'cfdi': 'http://www.sat.gob.mx/cfd/3',
    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
    'implocal': 'http://www.sat.gob.mx/implocal',
    'pago10': 'http://www.sat.gob.mx/Pagos',
}


class MainApplication(Frame):
    def __init__(self, master, *args, **kwargs):
        Frame.__init__(self, master, *args, **kwargs)

        self.master.title('CFDI 2 Excel')

        self.master.resizable(0, 0)

        self.lbl_titulo = Label(
            self, text='CFDI 2 Excel', font=('Arial Bold', 20)
        )
        self.lbl_titulo.grid(
            column=0, row=0, columnspan=4,
            sticky='ew', padx=10, pady=10
        )

        self.btn_fuente = Button(
            self, text='...', command=self.btn_fuente_click
        )
        self.btn_fuente.grid(row=1, column=0, sticky='ew', padx=10, pady=10)

        self.lbl_fuente = Label(self, text='Folder fuente:')
        self.lbl_fuente.grid(row=1, column=1, sticky='ew', padx=10, pady=10)

        self.lbl_path_fuente = Label(self, text=os.getcwd())
        self.lbl_path_fuente.grid(
            row=1, column=2, columnspan=2,
            sticky='ew', padx=10, pady=10
        )

        self.btn_destino = Button(
            self, text='...', command=self.btn_destino_click
        )
        self.btn_destino.grid(row=2, column=0, sticky='ew', padx=10, pady=10)

        self.lbl_destino = Label(self, text='Folder destino:')
        self.lbl_destino.grid(row=2, column=1, sticky='ew', padx=10, pady=10)

        self.lbl_path_destino = Label(self, text=os.getcwd())
        self.lbl_path_destino.grid(
            row=2, column=2, columnspan=2,
            sticky='ew', padx=10, pady=10
        )

        self.pgb_status = Progressbar(self)
        self.pgb_status.grid(
            row=3, column=0, columnspan=3,
            sticky='ew', padx=10, pady=10
        )

        self.btn_procesar = Button(
            self, text='Procesar', command=self.btn_procesar_click
        )
        self.btn_procesar.grid(row=3, column=3, sticky='ew', padx=10, pady=10)

        self.lbl_estado = Label(self, text='Listo')
        self.lbl_estado.grid(
            row=4, column=0, columnspan=4,
            sticky='ew', padx=10, pady=10
        )

    def btn_fuente_click(self):
        path = self.lbl_path_fuente['text']
        path = askdirectory(initialdir=path)
        if path:
            self.lbl_path_fuente['text'] = path

    def btn_destino_click(self):
        path = self.lbl_path_destino['text']
        path = askdirectory(initialdir=path)
        if path:
            self.lbl_path_destino['text'] = path

    def btn_procesar_click(self):
        res = askokcancel('Confirmar', '¿Seguro que quiere procesar?')

        if not res:
            return

        self.btn_fuente['state'] = 'disabled'
        self.btn_destino['state'] = 'disabled'
        self.btn_procesar['state'] = 'disabled'

        Thread(target=self.task_xml_to_excel).start()

    def task_xml_to_excel(self):
        self.pgb_status.start()

        fuente_path = self.lbl_path_fuente['text']
        destino_path = self.lbl_path_destino['text']

        files = [os.path.join(dp, f) for dp, dn, filenames in os.walk(
            fuente_path) for f in filenames if os.path.splitext(f)[1] == '.xml']

        text = 'XML encontrados: {}'.format(len(files))
        self.lbl_estado['text'] = text

        wb = Workbook()

        #sheet = wb.active
        cfdis_sheet = wb.create_sheet(title='CFDIs')
        pagos_sheet = wb.create_sheet(title="Pagos")

        cfdis_sheet.append(
            (
                'Version', 'UUID', 'Serie', 'Folio', 'Tipo', 'Fecha emision',
                'Fecha certificacion', 'pacCertifico', 'RFC emisor', 'Razon emisor',
                'RFC receptor', 'Razon receptor', 'Conceptos', 'Uso CFDI', 'Moneda',
                'Tipo de cambio', 'Metodo pago', 'Forma pago', 'SubTotal', 'Descuento',
                'IVA Trasladado', 'ISR Trasladado', 'IEPS Trasladado', 'IVA Retenido',
                'ISR Retenido', 'Impuesto Local', 'Total', 'TipoRelacion', 'CfdiRelacionados',
            )
        )
        pagos_sheet.append(
            (
                'Version', 'UUID', 'Serie', 'Folio', 'Tipo', 'Fecha emision',
                'Fecha certificacion', 'pacCertifico', 'RFC emisor', 'Razon emisor',
                'RFC receptor', 'Razon receptor', 'Conceptos', 'Uso CFDI', 'Moneda',
                'Tipo de cambio', 'Metodo pago', 'Forma pago', 'SubTotal', 'Descuento',
                'IVA Trasladado', 'ISR Trasladado', 'IEPS Trasladado', 'IVA Retenido',
                'ISR Retenido', 'Impuesto Local', 'Total', 'TipoRelacion', 'CfdiRelacionados',
                # pago10:Pago
                'FechaPago',
                'FormaDePagoP',
                'MonedaP',
                'Monto',
                'RfcEmisorCtaOrd',
                'NomBancoOrdExt',
                'CtaOrdenante',
                'RfcEmisorCtaBen',
                'CtaBeneficiario',
                'NumOperacion',
                # pago10:DoctoRelacionado
                'IdDocumento',
                'Folio',
                'MonedaDR',
                'MetodoDePagoDR',
                'NumParcialidad',
                'ImpSaldoAnt',
                'ImpPagado',
                'ImpSaldoInsoluto',
            )
        )

        for f in files:
            try:

                self.lbl_estado['text'] = 'Procesando: {}'.format(f)

                root = etree.parse(
                    f, parser=etree.XMLParser(huge_tree=True, recover=True)
                ).getroot()

                version = root.get('Version')

                uuid = root.find(
                    'cfdi:Complemento/tfd:TimbreFiscalDigital',
                    namespaces=NSMAP
                ).get('UUID')

                serie = root.get('Serie')

                folio = root.get('Folio')

                tipo = root.get('TipoDeComprobante')

                fecha = root.get('Fecha')

                fecha_timbrado = root.find(
                    'cfdi:Complemento/tfd:TimbreFiscalDigital',
                    namespaces=NSMAP
                ).get('FechaTimbrado')

                pac = root.find(
                    'cfdi:Complemento/tfd:TimbreFiscalDigital',
                    namespaces=NSMAP
                ).get('RfcProvCertif')

                rfc_emisor = root.find(
                    'cfdi:Emisor',
                    namespaces=NSMAP
                ).get('Rfc')

                nombre_emisor = root.find(
                    'cfdi:Emisor',
                    namespaces=NSMAP
                ).get('Nombre')

                rfc_receptor = root.find(
                    'cfdi:Receptor',
                    namespaces=NSMAP
                ).get('Rfc')

                nombre_receptor = root.find(
                    'cfdi:Receptor',
                    namespaces=NSMAP
                ).get('Nombre')

                conceptos = ''

                for i, c in enumerate(root.findall('cfdi:Conceptos/cfdi:Concepto', namespaces=NSMAP)):
                    conceptos += '|-{}-|: {}: {} '.format(
                        i + 1,
                        c.get('Descripcion'),
                        c.get('Importe')
                    )

                uso = root.find(
                    'cfdi:Receptor',
                    namespaces=NSMAP
                ).get('UsoCFDI')

                moneda = root.get('Moneda')

                tipo_cambio = root.get('TipoCambio')

                metodo_pago = root.get('MetodoPago')

                forma_pago = root.get('FormaPago')

                subtotal = root.get('SubTotal')

                descuento = root.get('Descuento')

                iva = 0.0
                isr = 0.0
                ieps = 0.0
                for t in root.findall('cfdi:Impuestos/cfdi:Traslados/cfdi:Traslado', namespaces=NSMAP):
                    if t.get('Impuesto') == '002':
                        iva += float(t.get('Importe'))
                    if t.get('Impuesto') == '001':
                        isr += float(t.get('Importe'))
                    if t.get('Impuesto') == '003':
                        ieps += float(t.get('Importe'))

                iva_ret = 0
                isr_ret = 0
                for t in root.findall('cfdi:Impuestos/cfdi:Retenciones/cfdi:Retencion', namespaces=NSMAP):
                    if t.get('Impuesto') == '002':
                        iva_ret += float(t.get('Importe'))
                    if t.get('Impuesto') == '001':
                        isr_ret += float(t.get('Importe'))

                total = root.get('Total')

                tipo_relacion = ''
                relaciones = ''

                cfdi_relacionados = root.find(
                    'cfdi:CfdiRelacionados', namespaces=NSMAP)

                if cfdi_relacionados is not None:

                    tipo_relacion = cfdi_relacionados.get('TipoRelacion')

                    for r in cfdi_relacionados.findall('cfdi:CfdiRelacionado', namespaces=NSMAP):
                        relaciones += '{}, '.format(
                            r.get('UUID')
                        )

                implocal = 0

                for t in root.findall('cfdi:Complemento/implocal:ImpuestosLocales/implocal:TrasladosLocales', namespaces=NSMAP):
                    implocal += float(t.get('Importe'))

                cfdis_sheet.append((
                    version, uuid, serie, folio, tipo, fecha,
                    fecha_timbrado, pac, rfc_emisor, nombre_emisor,
                    rfc_receptor, nombre_receptor, conceptos, uso, moneda,
                    tipo_cambio, metodo_pago, forma_pago, subtotal, descuento,
                    iva, isr, ieps, iva_ret, isr_ret, implocal, total, tipo_relacion, relaciones
                ))

                for docto_relacionado in root.findall('cfdi:Complemento/pago10:Pagos/pago10:Pago/pago10:DoctoRelacionado', namespaces=NSMAP):
                    pago = docto_relacionado.getparent()

                    pagos_sheet.append((
                        version, uuid, serie, folio, tipo, fecha,
                        fecha_timbrado, pac, rfc_emisor, nombre_emisor,
                        rfc_receptor, nombre_receptor, conceptos, uso, moneda,
                        tipo_cambio, metodo_pago, forma_pago, subtotal, descuento,
                        iva, isr, ieps, iva_ret, isr_ret, implocal, total, tipo_relacion, relaciones,
                        # pago10:Pago
                        pago.get('FechaPago'),
                        pago.get('FormaDePagoP'),
                        pago.get('MonedaP'),
                        pago.get('Monto'),
                        pago.get('RfcEmisorCtaOrd'),
                        pago.get('NomBancoOrdExt'),
                        pago.get('CtaOrdenante'),
                        pago.get('RfcEmisorCtaBen'),
                        pago.get('CtaBeneficiario'),
                        pago.get('NumOperacion'),
                        # pago10:DoctoRelacionado
                        docto_relacionado.get('IdDocumento'),
                        docto_relacionado.get('Folio'),
                        docto_relacionado.get('MonedaDR'),
                        docto_relacionado.get('MetodoDePagoDR'),
                        docto_relacionado.get('NumParcialidad'),
                        docto_relacionado.get('ImpSaldoAnt'),
                        docto_relacionado.get('ImpPagado'),
                        docto_relacionado.get('ImpSaldoInsoluto'),
                    ))

            except Exception as e:
                cfdis_sheet.append((str(e), ))

        file_path = os.path.join(destino_path, 'cfdis.xlsx')

        wb.save(file_path)

        self.pgb_status.stop()

        self.btn_fuente['state'] = 'normal'
        self.btn_destino['state'] = 'normal'
        self.btn_procesar['state'] = 'normal'

        showinfo(
            'Completado',
            'Proceso completado\nArchivo guardado en: {}'.format(file_path)
        )

        os.system('start excel "{}"'.format(file_path))


if __name__ == '__main__':
    root = Tk()

    MainApplication(root).grid(row=0, column=0)

    root.mainloop()
